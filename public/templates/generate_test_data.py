import os
import random
from datetime import datetime
import openpyxl

# ── Cấu hình ────────────────────────────────────────────────
TEMPLATE_FILE = "Report_Template.xlsx"
OUTPUT_DIR = "input"
NUM_FILES = 50
ROWS_PER_FILE = 2          # số dòng dữ liệu ghi vào mỗi file
FIRST_DATA_ROW = 4         # hàng 1 = tiêu đề, hàng 2 = nhóm cột, hàng 3 = tên cột, hàng 4 mới là dữ liệu thật

# ── Cột trong template (theo đúng thứ tự thật của file) ──────
# 1  Ngày
# 2  Ca
# 3  Đường lò
# 4  Đơn vị thi công
# 5  Người báo cáo
# 6  Số lao động ca
# 7  Sản lượng than thực hiện (tấn)
# 8  Tiết diện đào lò (m2)
# 9  Thực hiện đào lò (m)
# 10 Tiết diện xén lò (m2)
# 11 Thực hiện xén lò (m)
# 12 Bước chống (m/vì)
# 13 Thực hiện (m) - chống đội lò
# 14 Thực hiện khấu lò (m)
# 15 Lò nối: Thực hiện (m)
# 16 Tiến độ khấu chung: Thực hiện (m)
# 17 Giá chống hiện tại
# 18 Công việc khác
# 19 Ghi chú / Sự cố

# Danh sách dữ liệu mẫu để sinh ngẫu nhiên cho phong phú
don_vi_list = ["ĐL 1", "ĐL 2", "KT 1", "KT 2", "Cơ Điện"]
nguoi_bao_cao_list = ["Nguyễn Văn A", "Trần Văn B", "Lê Văn C", "Phạm Văn D", "Hoàng Thừa Vũ"]

duong_lo_list = [
    "DVVT LC 30708 đoạn A6-A8",
    "Lò chợ 30705",
    "XV lò TGVT -340/-275 khu III vỉa 7",
    "Lò dọc vỉa vận tải mức -50",
    "Thượng thông gió +30/-30",
]

cong_viec_khac_list = [
    "Kiểm tra thiết bị định kỳ",
    "Vệ sinh công nghiệp",
    "Tập huấn an toàn lao động",
    "Bảo dưỡng hệ thống thông gió",
]

# Kịch bản ghi chú/sự cố — chia rõ theo mức độ để kiểm soát tỉ lệ khi sinh dữ liệu
kich_ban_binh_thuong = [
    "Ổn định, tiến độ tốt.",
    "Bình thường, không có diễn biến lạ.",
    "Vận hành an toàn.",
    "Ổn định, diện khai thác khô ráo.",
    "Ổn định, sản xuất bình thường.",
]

kich_ban_canh_bao = [
    "Phát hiện lệch trục vì chống số 5, đã cho căn chỉnh lại.",
    "Áp lực địa tầng tăng cao tại lò TGVT, cần theo dõi sát.",
    "Thiếu hụt 3 lao động do xin nghỉ đột xuất ca 1.",
    "Xảy ra va chạm nhẹ khi vận hành xe gá, không có thiệt hại nhân mạng.",
    "Thiết bị máy khấu gặp sự cố kẹt xích ca 2, dừng sản xuất 1 tiếng.",
    "Khí metan có dấu hiệu tăng nhẹ 0.5%, hệ thống thông gió đã tăng công suất.",
    "Lò nứt nhẹ vì gỗ số 12, đã gia cố.",
    "Tiến độ chống đội chậm nghiêm trọng do nền lò lún bùn.",
    "Băng tải lò nối gặp sự cố rách bạt, ca trực đã vá xong.",
]

kich_ban_nghiem_trong = [
    "Sập lò cục bộ tại gương lò chợ 30705, đã tổ chức chống giữ xử lý xong.",
    "Xảy ra tai nạn lao động nhẹ, thợ lò bị dập ngón tay đã sơ cứu.",
]


# Tỉ lệ: ~60% bình thường, ~30% cảnh báo, ~10% nghiêm trọng — đủ đa dạng để test AI
def sinh_ghi_chu():
    r = random.random()
    if r < 0.60:
        return random.choice(kich_ban_binh_thuong)
    elif r < 0.90:
        return random.choice(kich_ban_canh_bao)
    else:
        return random.choice(kich_ban_nghiem_trong)


def clear_old_sample_rows(ws, from_row, to_row, max_col=19):
    """Xóa sạch dữ liệu mẫu có sẵn trong template (hàng 4-7) để không lẫn vào file test.
    Gán .value = None trực tiếp lên từng Cell (không dùng ws.cell(..., value=None),
    vì cách đó không xóa được giá trị đang có - openpyxl coi value=None là "không set gì").
    """
    for r in range(from_row, to_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).value = None


def main():
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Không tìm thấy file mẫu: {TEMPLATE_FILE}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("🚀 Đang tiến hành tạo file test dữ liệu...")

    for i in range(1, NUM_FILES + 1):
        wb = openpyxl.load_workbook(TEMPLATE_FILE)
        ws = wb.active

        # Xóa hết dữ liệu mẫu có sẵn từ hàng 4 trở đi trước khi ghi dữ liệu mới
        clear_old_sample_rows(ws, FIRST_DATA_ROW, ws.max_row)

        # Ngày + đường lò dùng chung cho các dòng trong cùng 1 file (đúng logic 1 file = 1 báo cáo ca/ngày)
        random_day = random.randint(1, 30)
        ngay_mock = datetime(2026, 7, random_day)
        ca_mock = random.choice([1, 2, 3])
        duong_lo_mock = random.choice(duong_lo_list)

        for offset in range(ROWS_PER_FILE):
            row = FIRST_DATA_ROW + offset
            ws.cell(row=row, column=1,  value=ngay_mock)                            # Ngày
            ws.cell(row=row, column=2,  value=ca_mock)                              # Ca
            ws.cell(row=row, column=3,  value=duong_lo_mock)                        # Đường lò
            ws.cell(row=row, column=4,  value=random.choice(don_vi_list))           # Đơn vị thi công
            ws.cell(row=row, column=5,  value=random.choice(nguoi_bao_cao_list))    # Người báo cáo
            ws.cell(row=row, column=6,  value=random.randint(15, 35))               # Số lao động ca
            ws.cell(row=row, column=7,  value=round(random.uniform(1.5, 6.0), 1))   # Sản lượng than thực hiện (tấn)
            ws.cell(row=row, column=8,  value=round(random.uniform(2.5, 4.0), 1))   # Tiết diện đào lò (m2)
            ws.cell(row=row, column=9,  value=round(random.uniform(1.0, 4.0), 1))   # Thực hiện đào lò (m)

            # Các hạng mục dưới đây không phải ca nào cũng có -> random để trống cho giống thực tế
            if random.random() < 0.7:
                ws.cell(row=row, column=10, value=round(random.uniform(2.0, 3.5), 1))   # Tiết diện xén lò (m2)
                ws.cell(row=row, column=11, value=round(random.uniform(0.5, 2.5), 1))   # Thực hiện xén lò (m)
            if random.random() < 0.7:
                ws.cell(row=row, column=12, value=random.choice([0.8, 1.0, 1.2]))       # Bước chống (m/vì)
                ws.cell(row=row, column=13, value=round(random.uniform(1.0, 3.0), 1))   # Thực hiện (m) - chống đội lò
            if random.random() < 0.6:
                ws.cell(row=row, column=14, value=random.randint(75, 95))               # Thực hiện khấu lò (m)
            if random.random() < 0.5:
                ws.cell(row=row, column=15, value=round(random.uniform(1.0, 3.5), 1))   # Lò nối: Thực hiện (m)

            ws.cell(row=row, column=16, value=round(random.uniform(1.5, 3.5), 1))   # Tiến độ khấu chung: Thực hiện (m)
            ws.cell(row=row, column=17, value=random.choice([18, 20, 21, 22]))      # Giá chống hiện tại
            if random.random() < 0.3:
                ws.cell(row=row, column=18, value=random.choice(cong_viec_khac_list))  # Công việc khác
            ws.cell(row=row, column=19, value=sinh_ghi_chu())                       # Ghi chú / Sự cố

        filename = f"bao_cao_test_{i:02d}.xlsx"
        file_output_path = os.path.join(OUTPUT_DIR, filename)
        try:
            wb.save(file_output_path)
        except PermissionError:
            print(f"⚠️  Bỏ qua '{filename}': file đang mở trong Excel, hãy đóng file đó rồi chạy lại script.")
        finally:
            wb.close()

    print(f"✔️ Thành công! Đã tạo xong {NUM_FILES} file báo cáo mẫu nằm gọn gàng trong thư mục: '{OUTPUT_DIR}/'")


if __name__ == "__main__":
    main()